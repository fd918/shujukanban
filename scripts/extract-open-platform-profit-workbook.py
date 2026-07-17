#!/usr/bin/env python3
"""把开放平台毛利 Excel 提取为前端构建脚本可读取的本地 JSON。"""

import json
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel


def json_value(value):
    if isinstance(value, (datetime, date)):
        return to_excel(value)
    return value


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("用法：extract-open-platform-profit-workbook.py <Excel路径> <JSON路径>")

    source_path = Path(sys.argv[1]).expanduser().resolve()
    output_path = Path(sys.argv[2]).expanduser().resolve()
    workbook = load_workbook(source_path, data_only=True, read_only=True)
    sheets = []

    for sheet in workbook.worksheets:
        values = [
            [json_value(value) for value in row]
            for row in sheet.iter_rows(min_row=1, max_row=23, min_col=1, max_col=35, values_only=True)
        ]
        sheets.append({
            "sheet": sheet.title,
            "range": "A1:AI23",
            "values": values,
            "formulas": [],
        })

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(sheets, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"已提取：{output_path}")


if __name__ == "__main__":
    main()
